from PyQt5 import QtCore, QtGui, QtWidgets, Qt
import traceback
import sys
from sort_and_filter_tbl_view import Sort_tbl_view
import pandas as pd
from mpl_cls import ScrollaleChartsArea
import os
import datetime
import numpy as np
import pandas.errors
import dateutil
from parameters import *
import pymongo
import db_operator
import collections
import pickle
import mpl_cls
import matplotlib.dates as mdates
import math
import functools
import openpyxl
from pandas import ExcelWriter
import time
import os
from shutil import copyfile

def on_pick(event, lined, fig):
    # On the pick event, find the original line corresponding to the legend
    # proxy line, and toggle its visibility.
    legline = event.artist
    origline = lined[legline]
    visible = not origline.get_visible()
    origline.set_visible(visible)
    # Change the alpha on the line in the legend so we can see what lines
    # have been toggled.
    legline.set_alpha(1.0 if visible else 0.2)
    fig.canvas.draw()


class ScrollaleChartsAreaInReportWidget(mpl_cls.ScrollaleChartsArea):
    def __init__(self, parent=None):
        super(ScrollaleChartsAreaInReportWidget, self).__init__(parent=parent)
        self.width = 800
        self.height = 600
        self.TITLE_PAD = 25
        self.recreate()

    def plot_2g(self, gsm_site_df, gsm_cell_df, config_df):
        for df in (gsm_site_df, gsm_cell_df):
            if df.empty:
                return
        gsm_site_df.sort_values(by="Date", inplace=True)
        gsm_config_df = config_df[(config_df["Tech"] == "2G") & (config_df["Chart_type"] == "Line")]
        df = gsm_site_df
        chart_no = 0
        site = gsm_site_df["Site Name"].unique()[0]
        for index, row in gsm_config_df.iterrows():
            plot_para = {"marker": "."}
            try:
                df[row["Kpi_name"]] = eval(row["Formula"])
            except (KeyError, SyntaxError, TypeError, ValueError, NameError) as e:
                print("wrong formula")
                print(row["Kpi_name"], row["Formula"])
                print(type(e), e)
                # traceback.print_exc()
                df[row["Kpi_name"]] = float('nan')
            widget = mpl_cls.MplWidget(parent=self.scroll_widget)
            self.widgets[chart_no,0] = widget
            self.gridlayout.addWidget(widget, chart_no, 0)
            widget.setFixedWidth(self.width)
            widget.setFixedHeight(self.height)
            ax = widget.canvas.fig.add_subplot(111)
            ax.ticklabel_format(style='plain', useOffset=False)
            ax.grid(axis='y')
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%d.%m'))

            x = df["Date"]
            y = df[row["Kpi_name"]]
            ax.plot(x, y, **plot_para)
            chart_no += 1
            ax.set_title("{} , {}".format(row["Kpi_name"], site), pad=self.TITLE_PAD)
            widget.canvas.fig.autofmt_xdate()

        gsm_cell_df.sort_values(by=["Cell Name", "Date"], inplace=True)
        df = gsm_cell_df
        chart_no = 0
        for index, row in gsm_config_df.iterrows():
            lined = {}
            lines = []
            plot_para = {"marker": "."}
            try:
                df[row["Kpi_name"]] = eval(row["Formula"])
            except (KeyError, SyntaxError, TypeError, ValueError, NameError) as e:
                print("wrong formula")
                print(row["Kpi_name"], row["Formula"])
                print(type(e), e)
                # traceback.print_exc()
                df[row["Kpi_name"]] = float('nan')
            widget = mpl_cls.MplWidget(parent=self.scroll_widget)
            fig = widget.canvas.fig
            self.widgets[chart_no,1] = widget
            self.gridlayout.addWidget(widget, chart_no, 1)
            widget.setFixedWidth(self.width)
            widget.setFixedHeight(self.height)
            ax = widget.canvas.fig.add_subplot(111)
            ax.ticklabel_format(style='plain', useOffset=False)
            ax.grid(axis='y')
            cell_count = 0
            for cell in df["Cell Name"].unique():
                cell_df = df[df["Cell Name"] == cell]
                x = cell_df["Date"]
                y = cell_df[row["Kpi_name"]]
                plot_para['label'] = cell
                plot_para["color"] = "C" + str(cell_count)
                line, = ax.plot(x, y, **plot_para)
                lines.append(line)
                cell_count += 1
            ax.set_title("{} , {}".format(row["Kpi_name"], site), pad=self.TITLE_PAD)
            legend_para = {"loc": "upper left", "bbox_to_anchor": (-0.1, 1.12), "frameon": False, "fontsize": 10}
            legend_para["ncol"] = cell_count
            if cell_count >= 4:
                legend_para["fontsize"] = 6
                legend_para["ncol"] = math.ceil(cell_count / 2)
                legend_para["columnspacing"] = 0.2
                ax.set_title("{} , {}".format(row["Kpi_name"], site), pad=self.TITLE_PAD)

            leg = ax.legend(**legend_para)
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%d.%m'))
            widget.canvas.fig.autofmt_xdate()
            for legline, origline in zip(leg.get_lines(), lines):
                legline.set_picker(5)  # Enable picking on the legend line.
                lined[legline] = origline
            fig.canvas.mpl_connect('pick_event', functools.partial(on_pick, lined=lined, fig=fig))
            chart_no += 1

    def plot_4g(self, lte_site_df, lte_cell_df, config_df):
        for df in (lte_site_df, lte_cell_df):
            if df.empty:
                return
        lte_site_df.sort_values(by="Date", inplace=True)
        lte_config_df = config_df[(config_df["Tech"] == "4G") & (config_df["Chart_type"] == "Line")]
        df = lte_site_df
        chart_no = 0
        site = lte_site_df["eNodeB Name"].unique()[0]
        for index, row in lte_config_df.iterrows():
            plot_para = {"marker": "."}
            try:
                df[row["Kpi_name"]] = eval(row["Formula"])
            except (KeyError, SyntaxError, TypeError, ValueError, NameError) as e:
                print("wrong formula")
                print(row["Kpi_name"], row["Formula"])
                print(type(e), e)
                # traceback.print_exc()
                df[row["Kpi_name"]] = float('nan')
            widget = mpl_cls.MplWidget(parent=self.scroll_widget)
            self.widgets[chart_no,0] = widget
            self.gridlayout.addWidget(widget, chart_no, 0)
            widget.setFixedWidth(self.width)
            widget.setFixedHeight(self.height)
            ax = widget.canvas.fig.add_subplot(111)
            ax.ticklabel_format(style='plain', useOffset=False)
            ax.grid(axis='y')
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%d.%m'))

            x = df["Date"]
            y = df[row["Kpi_name"]]
            ax.plot(x, y, **plot_para)
            chart_no += 1
            ax.set_title("{} , {}".format(row["Kpi_name"], site), pad=self.TITLE_PAD)
            widget.canvas.fig.autofmt_xdate()

        lte_cell_df.sort_values(by=["Cell Name", "Date"], inplace=True)
        df = lte_cell_df
        chart_no = 0
        for index, row in lte_config_df.iterrows():
            lined = {}
            lines = []
            plot_para = {"marker": "."}
            try:
                df[row["Kpi_name"]] = eval(row["Formula"])
            except (KeyError, SyntaxError, TypeError, ValueError, NameError) as e:
                print("wrong formula")
                print(row["Kpi_name"], row["Formula"])
                print(type(e), e)
                # traceback.print_exc()
                df[row["Kpi_name"]] = float('nan')
            widget = mpl_cls.MplWidget(parent=self.scroll_widget)
            fig = widget.canvas.fig
            self.widgets[chart_no,1] = widget
            self.gridlayout.addWidget(widget, chart_no, 1)
            widget.setFixedWidth(self.width)
            widget.setFixedHeight(self.height)
            ax = widget.canvas.fig.add_subplot(111)
            ax.ticklabel_format(style='plain', useOffset=False)
            ax.grid(axis='y')
            cell_count = 0
            for cell in df["Cell Name"].unique():
                cell_df = df[df["Cell Name"] == cell]
                x = cell_df["Date"]
                y = cell_df[row["Kpi_name"]]
                plot_para['label'] = cell
                plot_para["color"] = "C" + str(cell_count)
                line, = ax.plot(x, y, **plot_para)
                lines.append(line)
                cell_count += 1
            legend_para = {"loc": "upper left", "bbox_to_anchor": (-0.1, 1.1), "frameon": False, "fontsize": 10}
            legend_para["ncol"] = cell_count
            ax.set_title("{} , {}".format(row["Kpi_name"], site), pad=self.TITLE_PAD)
            if cell_count >= 4:
                legend_para["fontsize"] = 6
                legend_para["ncol"] = math.ceil(cell_count / 2)
                legend_para["columnspacing"] = 0.2
                legend_para["bbox_to_anchor"] = (-0.1, 1.1)
                ax.set_title("{} , {}".format(row["Kpi_name"], site), pad=self.TITLE_PAD)

            leg = ax.legend(**legend_para)
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%d.%m'))
            widget.canvas.fig.autofmt_xdate()
            for legline, origline in zip(leg.get_lines(), lines):
                legline.set_picker(5)  # Enable picking on the legend line.
                lined[legline] = origline
            fig.canvas.mpl_connect('pick_event', functools.partial(on_pick, lined=lined, fig=fig))
            chart_no += 1


class Report_widget(QtWidgets.QWidget):
    def __init__(self, parent=None):
        super(Report_widget, self).__init__(parent)
        self.splitter = QtWidgets.QSplitter(self)
        self.splitter.setObjectName(u"splitter")
        self.splitter.setOrientation(QtCore.Qt.Horizontal)
        self.left_list = QtWidgets.QListWidget(self.splitter)
        self.left_list.addItem("Input")
        self.left_list.setObjectName(u"list_view")
        self.splitter.addWidget(self.left_list)
        self.stacked_widget = QtWidgets.QStackedWidget(self.splitter)
        self.stacked_widget.setObjectName(u"stacked_widget")
        self.input_page = QtWidgets.QWidget(self)
        self.input_page.setObjectName(u"input_page")
        self.input_report_btn = QtWidgets.QPushButton(parent=self.input_page, text="Load Report Input")
        self.export_single_site_btn = QtWidgets.QPushButton(parent=self.input_page, text="Export Single Site Report")
        self.export_all_site_btn = QtWidgets.QPushButton(parent=self.input_page, text="Export All Sites Reports")
        input_page_gridlayout = QtWidgets.QGridLayout(self.input_page)
        input_page_gridlayout.addWidget(self.input_report_btn, 0, 0)
        input_page_gridlayout.addWidget(self.export_single_site_btn, 0, 1)
        input_page_gridlayout.addWidget(self.export_all_site_btn, 0, 2)
        self.report_input_tbl = Sort_tbl_view(self.input_page)
        input_page_gridlayout.addWidget(self.report_input_tbl, 1, 0, 1, -1)
        self.input_page.setLayout(input_page_gridlayout)
        self.stacked_widget.addWidget(self.input_page)
        self.splitter.addWidget(self.stacked_widget)
        layout = QtWidgets.QHBoxLayout(self)
        layout.addWidget(self.splitter)
        self.setLayout(layout)
        self.stacked_widgets = {}
        self.connect()

    def connect(self):
        self.left_list.currentRowChanged.connect(self.on_click_left_list_currentRowChanged)
        self.input_report_btn.clicked.connect(self.on_input_report_btn_clicked)
        self.export_single_site_btn.clicked.connect(self.on_export_single_site_btn_clicked)

    def set_input_page_df(self, df):
        self.report_input_tbl.set_df(df)
        # print(self.sites_list_tbl.)

    def on_click_left_list_currentRowChanged(self, i):
        self.stacked_widget.setCurrentIndex(i)

    def on_input_report_btn_clicked(self):
        fn = QtWidgets.QFileDialog.getOpenFileName(self, "Open ppo report input file", os.getcwd(), "*.xlsx")[0]
        if fn != "":
            valid, df = self.validate_report_input(fn)
            if valid:
                self.set_input_page_df(df)
            else:
                err = df
                QtWidgets.QErrorMessage(self).showMessage(err)

    def validate_report_input(self, fn):
        pass

    def on_export_single_site_btn_clicked(self):
        pass

    def on_export_all_site_btn_clicked(self):
        pass

    def add_tbl_tab(self, title, df):
        tbl_widget = Sort_tbl_view(self.stacked_widget)
        tbl_widget.set_df(df)
        self.left_list.addItem(title)
        self.stacked_widget.addWidget(tbl_widget)
        self.stacked_widgets[title] = tbl_widget

    def add_charts_tab(self, title, sc):
        self.left_list.addItem(title)
        self.stacked_widget.addWidget(sc)
        self.stacked_widgets[title] = sc

    def clear_tabs(self):
        for title, w in self.stacked_widgets.items():
            list_widget = self.left_list.findItems(title, QtCore.Qt.MatchExactly)[0]
            self.left_list.takeItem(self.left_list.indexFromItem(list_widget).row())
            self.stacked_widget.removeWidget(w)
            w.deleteLater()
            del list_widget
        self.stacked_widgets = {}


class Ppo_report_widget(Report_widget):
    def __init__(self, config_df, query_config, agg_function, addtional_kpi, conditional_kpi, ignore_fields,
                 parent=None):
        super(Ppo_report_widget, self).__init__(parent=parent)
        self.config_df = config_df
        self.query_config = query_config
        self.agg_function = agg_function
        self.additional_kpi = addtional_kpi
        self.conditional_kpi = conditional_kpi
        self.ignore_fields = ignore_fields
        self.input_columns = {"DU ID": 'O', "Pre Start": 'datetime64[ns]', "Pre End": 'datetime64[ns]',
                              "Post Start": 'datetime64[ns]', "Post End": 'datetime64[ns]', "Days": 'int64',
                              "Day1": 'datetime64[ns]', "Day2": 'datetime64[ns]'}
        self.set_input_page_df(pd.DataFrame(columns=self.input_columns))
        self.filling_data = {"Site ID": ("Overview", "G4"),
                             "DU ID": ("Overview", "G5"),
                             "Site Name": ("Overview", "G6"),
                             "NR BIS": ("Overview", "G12"),
                             "E2E-MS12A: 4G SITE COMMERCIALISED (BIS)": ("Overview", "G13")
                             }
        project = set()
        for kpis in self.config_df["Kpis"]:
            for kpi in kpis:
                project.add(kpi)
        self.project = project

    def validate_report_input(self, fn):
        df = pd.read_excel(fn)
        if list(df.columns) != list(self.input_columns.keys()):
            err = "Wrong title, title must be the same."
            return False, err
        else:
            try:
                df = df.astype(self.input_columns)
                if df.isnull().values.any():
                    err = "empty values in input file"
                    return False, err
            except (dateutil.parser._parser.ParserError, ValueError) as e:
                print("values wrong in input file")
                print(type(e), e)
                err = "values wrong in input file"
                return False, err
            else:
                for ind, row in df.iterrows():
                    if not (datetime.datetime.today() - datetime.timedelta(days=row["Days"]) <= row[
                        "Day1"] <= datetime.datetime.today()) or not (
                            datetime.datetime.today() - datetime.timedelta(days=row["Days"]) <= row[
                        "Day2"] <= datetime.datetime.today()):
                        err = "row {} : Day1 or Day2 not in range".format(ind)
                        return False, err

                return True, df

    def on_export_single_site_btn_clicked(self):
        self.clear_tabs()
        df = self.report_input_tbl.model.df
        row = self.report_input_tbl.currentIndex().row()
        if row >= 0:
            input_df = df.loc[row:row]
            # get site and from isdp
            input_df, missing_site = self.get_isdp_info(input_df)
            if missing_site:
                QtWidgets.QErrorMessage(self).showMessage(
                    "Cannot find below DU_ID in ISDP: \n {}".format(" ".join(missing_site)))
            for index, row in input_df.iterrows():
                self.export_ppo_report(row, in_gui=True)

    def get_isdp_info(self, input_df):
        myclient = pymongo.MongoClient(MONGO_CLIENT_URL)
        mydb = myclient[EP_DB_NAME]
        mycol = mydb[ISDP_COL_NAME]
        arr = input_df.to_dict("records")
        missing_duid = []
        for row in arr:
            du_id = row["DU ID"]
            d = mycol.find_one({"_id": du_id})
            if d:
                for k in self.filling_data:
                    if k in d:
                        row[k] = d[k]
            else:
                missing_duid.append(du_id)
        myclient.close()
        return pd.DataFrame(arr), missing_duid

    def get_tbl_data(self, input_df_row):
        pre_st, pre_end = input_df_row["Pre Start"], input_df_row["Pre End"]
        post_st, post_end = input_df_row["Post Start"], input_df_row["Post End"]
        sites_list = [input_df_row.get("Site ID")] if input_df_row.get("Site ID") else []

        if sites_list:
            # 2G table
            config_df = self.config_df[(self.config_df["Tech"] == "2G") & (self.config_df["Chart_type"] == "table")]
            dfs = []
            for st, end in ((pre_st, pre_end), (post_st, post_end)):
                data_collector_2g = db_operator.Data_collector(MONGO_CLIENT_URL=MONGO_CLIENT_URL, tech='2G',
                                                               agg_level="Overall", time_level="Daily",
                                                               agg_function=self.agg_function,
                                                               additional_kpi=self.additional_kpi,
                                                               conditional_kpi=self.conditional_kpi,
                                                               start_time=st, end_time=end,
                                                               project=self.project,
                                                               key_values=sites_list, query_col="Site Name",
                                                               query_config=self.query_config,
                                                               layer=None, cluster=None,
                                                               ignore_fields=self.ignore_fields
                                                               )
                data_collector_2g.query_chart_data()
                df = data_collector_2g.final_df
                for index, row in config_df.iterrows():
                    try:
                        df[row["Kpi_name"]] = eval(row["Formula"])
                    except (KeyError, SyntaxError, TypeError, ValueError, NameError) as e:
                        print("wrong formula")
                        print(row["Kpi_name"], row["Formula"])
                        print(type(e), e)
                        # traceback.print_exc()
                        df[row["Kpi_name"]] = float('nan')
                df = df[config_df["Kpi_name"]]
                dfs.append(df)
            pre_series = dfs[0].iloc[0]
            post_series = dfs[1].iloc[0]
            status_series = config_df[["Kpi_name", "Delta lte 0 Pass"]].set_index("Kpi_name")
            df = pd.concat([pre_series, post_series, status_series], axis=1)
            df.index.rename("KPI", inplace=True)
            df.columns = ["Pre", "Post", "Status"]
            df["Delta"] = (df["Post"] - df["Pre"]) / df["Pre"] * 100
            df["Status"] = df.apply(
                lambda row: None if pd.isnull(row["Delta"]) else (
                    "Pass" if (row["Delta"] >= 0) == row["Status"] else "Fail"),
                axis=1)
            df.reset_index(inplace=True)
            df = df.reindex(["KPI", "Pre", "Post", "Delta", "Status"], axis=1)
            gsm_tbl = df

            # lte_tbl
            config_df = self.config_df[(self.config_df["Tech"] == "4G") & (self.config_df["Chart_type"] == "table")]
            dfs = []

            for col, st, end in (("Pre", pre_st, pre_end), ("Post", post_st, post_end)):
                data_collector_4g = db_operator.Data_collector(MONGO_CLIENT_URL=MONGO_CLIENT_URL, tech='4G',
                                                               agg_level="Overall", time_level="Daily",
                                                               agg_function=self.agg_function,
                                                               additional_kpi=self.additional_kpi,
                                                               conditional_kpi=self.conditional_kpi,
                                                               start_time=st, end_time=end,
                                                               project=self.project,
                                                               key_values=sites_list,
                                                               query_col="eNodeB Name",
                                                               query_config=self.query_config,
                                                               layer=None, cluster=None,
                                                               ignore_fields=self.ignore_fields
                                                               )
                data_collector_4g.query_chart_data()
                df = data_collector_4g.final_df
                for index, row in config_df.iterrows():
                    try:
                        df[row["Kpi_name"]] = eval(row["Formula"])
                    except (KeyError, SyntaxError, TypeError, ValueError, NameError) as e:
                        print("wrong formula")
                        print(row["Kpi_name"], row["Formula"])
                        print(type(e), e)
                        # traceback.print_exc()
                        df[row["Kpi_name"]] = float('nan')
                df = df[config_df["Kpi_name"]]
                dfs.append(df)
            pre_series = dfs[0].iloc[0]
            post_series = dfs[1].iloc[0]
            status_series = config_df[["Kpi_name", "Delta lte 0 Pass"]].set_index("Kpi_name")
            df = pd.concat([pre_series, post_series, status_series], axis=1)
            df.index.rename("KPI", inplace=True)
            df.columns = ["Pre", "Post", "Status"]
            df["Delta"] = (df["Post"] - df["Pre"]) / df["Pre"] * 100
            df["Status"] = df.apply(
                lambda row: None if pd.isnull(row["Delta"]) else (
                    "Pass" if (row["Delta"] >= 0) == row["Status"] else "Fail"),
                axis=1)
            df.reset_index(inplace=True)
            df = df.reindex(["KPI", "Pre", "Post", "Delta", "Status"], axis=1)
            lte_tbl = df

            lte_layer_tbl = {}
            dfs = []
            for st, end in ((pre_st, pre_end), (post_st, post_end)):
                data_collector_4g_cell_level = db_operator.Data_collector(MONGO_CLIENT_URL=MONGO_CLIENT_URL,
                                                                          tech='4G',
                                                                          agg_level="Cell",
                                                                          time_level="Daily",
                                                                          agg_function=self.agg_function,
                                                                          additional_kpi=self.additional_kpi,
                                                                          conditional_kpi=self.conditional_kpi,
                                                                          start_time=st, end_time=end,
                                                                          project=self.project,
                                                                          key_values=sites_list,
                                                                          query_col="eNodeB Name",
                                                                          query_config=self.query_config,
                                                                          layer=None, cluster=None,
                                                                          ignore_fields=self.ignore_fields,
                                                                          )
                data_collector_4g_cell_level.query_chart_data()
                df = data_collector_4g_cell_level.final_df
                agg_function = {}
                for k in df.columns:
                    if k not in self.ignore_fields:
                        if k in self.agg_function:
                            agg_function[k] = self.agg_function[k]
                        else:
                            agg_function[k] = lambda x: x.sum(min_count=1)
                if not df.empty:
                    df = df.groupby(["layer", "Cell Name"], dropna=False).agg(agg_function)
                    df.sort_values(by=["layer", "Cell Name"], inplace=True)

                for index, row in config_df.iterrows():
                    try:
                        df[row["Kpi_name"]] = eval(row["Formula"])
                    except (KeyError, SyntaxError, TypeError, ValueError, NameError) as e:
                        print("wrong formula")
                        print(row["Kpi_name"], row["Formula"])
                        print(type(e), e)
                        # traceback.print_exc()
                        df[row["Kpi_name"]] = float('nan')
                df = df[config_df["Kpi_name"]]
                dfs.append(df)

            d = {}
            for index, row in dfs[0].iterrows():
                for col in row.index:
                    d[tuple(list(index) + [col])] = row[col]
            pre_series = pd.Series(d)
            d = {}
            for index, row in dfs[1].iterrows():
                for col in row.index:
                    d[tuple(list(index) + [col])] = row[col]
            post_series = pd.Series(d)
            status_series = config_df[["Kpi_name", "Delta lte 0 Pass"]].set_index("Kpi_name")["Delta lte 0 Pass"]
            df = pd.concat([pre_series, post_series], axis=1)
            df.columns = ["Pre", "Post"]
            df["Status"] = df.apply(lambda row: status_series.loc[row.name[2]], axis=1)
            df.index.rename(["layer", "Cell", "KPI"], inplace=True)
            df["Delta"] = (df["Post"] - df["Pre"]) / df["Pre"] * 100
            df["Status"] = df.apply(
                lambda row: None if pd.isnull(row["Delta"]) else (
                    "Pass" if (row["Delta"] >= 0) == row["Status"] else "Fail"),
                axis=1)
            df.reset_index(level=["Cell", "KPI"], inplace=True)
            df = df.reindex(["Cell", "KPI", "Pre", "Post", "Delta", "Status"], axis=1)

            for layer in df.index.unique():
                lte_layer_tbl[layer] = df.loc[layer]
            return gsm_tbl, lte_tbl, lte_layer_tbl

    def get_chart_data(self, input_df_row):
        st, end = datetime.datetime.today() - datetime.timedelta(
            days=int(input_df_row["Days"])), datetime.datetime.today()
        # st, end = datetime.datetime(2020, 1, 23), datetime.datetime(2020, 7, 29)

        sites_list = [input_df_row.get("Site ID")] if input_df_row.get("Site ID") else []
        if sites_list:
            # 2G table
            data_collector_2g_site = db_operator.Data_collector(MONGO_CLIENT_URL=MONGO_CLIENT_URL, tech='2G',
                                                                agg_level="Site", time_level="Daily",
                                                                agg_function=self.agg_function,
                                                                additional_kpi=self.additional_kpi,
                                                                conditional_kpi=self.conditional_kpi,
                                                                start_time=st, end_time=end,
                                                                project=self.project,
                                                                key_values=sites_list, query_col="Site Name",
                                                                query_config=self.query_config,
                                                                layer=None, cluster=None,
                                                                ignore_fields=self.ignore_fields
                                                                )
            data_collector_2g_site.query_chart_data()
            gsm_site_df = data_collector_2g_site.final_df
            data_collector_2g_cell = db_operator.Data_collector(MONGO_CLIENT_URL=MONGO_CLIENT_URL, tech='2G',
                                                                agg_level="Cell", time_level="Daily",
                                                                agg_function=self.agg_function,
                                                                additional_kpi=self.additional_kpi,
                                                                conditional_kpi=self.conditional_kpi,
                                                                start_time=st, end_time=end,
                                                                project=self.project,
                                                                key_values=sites_list, query_col="Site Name",
                                                                query_config=self.query_config,
                                                                layer=None, cluster=None,
                                                                ignore_fields=self.ignore_fields
                                                                )
            data_collector_2g_cell.query_chart_data()
            gsm_cell_df = data_collector_2g_cell.final_df

            data_collector_4g_site = db_operator.Data_collector(MONGO_CLIENT_URL=MONGO_CLIENT_URL, tech='4G',
                                                                agg_level="Site", time_level="Daily",
                                                                agg_function=self.agg_function,
                                                                additional_kpi=self.additional_kpi,
                                                                conditional_kpi=self.conditional_kpi,
                                                                start_time=st, end_time=end,
                                                                project=self.project,
                                                                key_values=sites_list, query_col="eNodeB Name",
                                                                query_config=self.query_config,
                                                                layer="All Site", cluster=None,
                                                                ignore_fields=self.ignore_fields
                                                                )
            data_collector_4g_site.query_chart_data()
            lte_site_df = data_collector_4g_site.final_df
            data_collector_4g_cell = db_operator.Data_collector(MONGO_CLIENT_URL=MONGO_CLIENT_URL, tech='4G',
                                                                agg_level="Cell", time_level="Daily",
                                                                agg_function=self.agg_function,
                                                                additional_kpi=self.additional_kpi,
                                                                conditional_kpi=self.conditional_kpi,
                                                                start_time=st, end_time=end,
                                                                project=self.project,
                                                                key_values=sites_list, query_col="eNodeB Name",
                                                                query_config=self.query_config,
                                                                layer=None, cluster=None,
                                                                ignore_fields=self.ignore_fields
                                                                )
            data_collector_4g_cell.query_chart_data()
            lte_cell_df = data_collector_4g_cell.final_df
            return gsm_site_df, gsm_cell_df, lte_site_df, lte_cell_df

    def show_tbl_data(self, gsm_tbl, lte_tbl, lte_layer_tbl):
        self.add_tbl_tab("2G", gsm_tbl)
        self.add_tbl_tab("4G", lte_tbl)
        for layer in sorted(lte_layer_tbl):
            self.add_tbl_tab(layer, lte_layer_tbl[layer])

    def export_ppo_report(self, row, in_gui):
        gsm_tbl, lte_tbl, lte_layer_tbl = self.get_tbl_data(row)
        if in_gui:
            self.show_tbl_data(gsm_tbl, lte_tbl, lte_layer_tbl)
        gsm_site_df, gsm_cell_df, lte_site_df, lte_cell_df = self.get_chart_data(row)
        parent = self.stacked_widget if in_gui else None
        sc_2g = ScrollaleChartsAreaInReportWidget(parent)
        sc_2g.plot_2g(gsm_site_df, gsm_cell_df, self.config_df)

        sc_4g = ScrollaleChartsAreaInReportWidget(parent)
        sc_4g.plot_4g(lte_site_df, lte_cell_df, self.config_df)

        self.export_to_excel(row, gsm_tbl, lte_tbl, lte_layer_tbl, sc_2g, sc_4g)
        if in_gui:
            self.add_charts_tab("2G Charts", sc_2g)
            self.add_charts_tab("4G Charts", sc_4g)

    def export_to_excel(self, row, gsm_tbl, lte_tbl, lte_layer_tbl, sc_2g, sc_4g,
                        template_fn=os.path.join("templates","ppm_report_templ.xlsx")):
        # copyfile(template_fn,fn)
        fn=os.path.join("ppo_reports","{}_ppm_report.xlsx".format(row["DU ID"]))
        wb = openpyxl.load_workbook(template_fn)

        for k, v in self.filling_data.items():
            ws, cell_no = v
            wb[ws][cell_no].value = row[k]
        wb['Overview']["G16"].value=row["Pre Start"].strftime("%d.%m.%Y")+" - "+row["Pre End"].strftime("%d.%m.%Y")
        wb['Overview']["G17"].value=row["Post Start"].strftime("%d.%m.%Y")+" - "+row["Post End"].strftime("%d.%m.%Y")
        wb.save(fn)

        writer = pd.ExcelWriter(fn,engine="openpyxl",modes="a")
        writer.book=wb
        gsm_tbl.to_excel(writer, sheet_name ='2G',index=False)
        lte_tbl.to_excel(writer, sheet_name ='4G',index=False)
        for layer in sorted(lte_layer_tbl):
            lte_layer_tbl[layer].to_excel(writer, sheet_name =layer.replace("/","-"),index=False)
        writer.save()
        writer.close()

        wb = openpyxl.load_workbook(fn)
        ws=wb.create_sheet(title="2G Charts")
        for k in sorted(sc_2g.widgets):
            r,c=k
            if c==0:
                c="A"
            else:
                c="L"
            r=1+26*r
            widget=sc_2g.widgets[k]
            fig=widget.canvas.fig
            s=str(time.time()).replace(".","")
            fig.savefig("temp/{}.png".format(s),dpi=100)
            img = openpyxl.drawing.image.Image("temp/{}.png".format(s))
            ws.add_image(img,c+str(r))

        ws=wb.create_sheet(title="4G Charts")
        for k in sorted(sc_4g.widgets):
            r,c=k
            if c==0:
                c="A"
            else:
                c="L"
            r=1+26*r
            widget=sc_4g.widgets[k]
            fig=widget.canvas.fig
            s=str(time.time()).replace(".","")
            fig.savefig("temp/{}.png".format(s),dpi=100)
            img = openpyxl.drawing.image.Image("temp/{}.png".format(s))
            ws.add_image(img,c+str(r))

        wb.save(fn)

# if __name__ == "__main__":
#     def excepthook(exc_type, exc_value, exc_tb):
#         tb = "".join(traceback.format_exception(exc_type, exc_value, exc_tb))
#         print("error catched!:")
#         print("error message:\n", tb)
#         QtWidgets.QApplication.quit()
#         # or QtWidgets.QApplication.exit(0)
#
#
#     sys.excepthook = excepthook
#     app = QtWidgets.QApplication([])
#     config_df = pd.read_excel("template.xlsx", sheet_name="ppo_config")
#     myWin = Ppo_report_widget()
#     myWin.show()
#     sys.exit(app.exec_())
